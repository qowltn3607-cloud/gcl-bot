from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

EXCEL_PATH = os.getenv("EXCEL_PATH", "GCL_추가견적_LOG.xlsx")
KRW_RATE   = int(os.getenv("KRW_RATE", 1380))

SAMSUNG_BLUE = "1428A0"
ROW_EVEN     = "F5F7FF"
WHITE        = "FFFFFF"

def _border():
    s = Side(border_style="thin", color="C5CAE9")
    return Border(left=s, right=s, top=s, bottom=s)

def _style_row(ws, row_num, is_even):
    bg = ROW_EVEN if is_even else WHITE
    aligns = {
        "A": "center", "B": "center", "C": "center",
        "D": "left",   "E": "right"
    }
    for col in ["A", "B", "C", "D", "E"]:
        c = ws[f"{col}{row_num}"]
        c.fill      = PatternFill("solid", fgColor=bg)
        c.font      = Font(name="Arial", size=9)
        c.border    = _border()
        c.alignment = Alignment(
            horizontal=aligns[col],
            vertical="center",
            wrap_text=(col == "D")
        )
    ws[f"E{row_num}"].number_format = '#,##0.00'

def _update_total(ws):
    # 마지막 데이터 행 찾기
    last_data_row = 2
    for r in range(3, ws.max_row + 1):
        if ws[f"A{r}"].value not in (None, "TOTAL"):
            last_data_row = r

    total_row = last_data_row + 2

    # 기존 TOTAL 행 삭제 후 재작성
    for r in range(3, ws.max_row + 2):
        if ws[f"A{r}"].value == "TOTAL":
            ws.delete_rows(r)
            break

    ws.row_dimensions[total_row].height = 28
    for col in ["A", "B", "C", "D"]:
        c = ws[f"{col}{total_row}"]
        c.fill   = PatternFill("solid", fgColor=SAMSUNG_BLUE)
        c.border = _border()
    ws[f"A{total_row}"].value     = "TOTAL"
    ws[f"A{total_row}"].font      = Font(name="Arial", bold=True, color=WHITE, size=10)
    ws[f"A{total_row}"].alignment = Alignment(horizontal="right", vertical="center")

    tc = ws[f"E{total_row}"]
    tc.value          = f"=SUM(E3:E{last_data_row})"
    tc.number_format  = '#,##0.00'
    tc.font           = Font(name="Arial", bold=True, color=WHITE, size=10)
    tc.fill           = PatternFill("solid", fgColor=SAMSUNG_BLUE)
    tc.alignment      = Alignment(horizontal="right", vertical="center")
    tc.border         = _border()

def append_estimate(sender: str, message: str, amount: float) -> int:
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    # 다음 데이터 행 & NO 계산
    next_row = 3
    for r in range(3, ws.max_row + 2):
        val = ws[f"A{r}"].value
        if val is None or val == "TOTAL":
            next_row = r
            break

    no       = next_row - 2
    is_even  = (no % 2 == 0)

    ws.row_dimensions[next_row].height = 22
    ws[f"A{next_row}"].value = no
    ws[f"B{next_row}"].value = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws[f"C{next_row}"].value = sender
    ws[f"D{next_row}"].value = message
    ws[f"E{next_row}"].value = amount

    _style_row(ws, next_row, is_even)
    _update_total(ws)

    wb.save(EXCEL_PATH)
    return no
